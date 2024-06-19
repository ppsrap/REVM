import datetime
import os
import openpyxl
import re
import requests

from gotoassist.setting import Auth_token_input
from gotoassist.transToCustomerMap import id_email_map
import logging

from gotoassist.transToServiceMap import servicesName_id_map

# set log level and format
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def getNewRows(folder_path):
    # Read the latest ID from the latestID.txt file
    rows = []  # Define rows here
    latest_id_file = os.path.join(folder_path, 'latestID.txt')
    with open(latest_id_file, 'r') as file:
        latest_id = int(file.read().strip())

    # Get all Excel files in the folder
    excel_files = [f for f in os.listdir(folder_path) if
                   f.startswith('Smartshares Incident_Request') and f.endswith('.xlsx')]
    max_file = None
    max_number = latest_id

    for file in excel_files:
        try:
            # Use regular expression to match the number after "-"
            match = re.search(r'-(\d+)', file)
            if match:
                # Extract the number and print it
                number = int(match.group(1))  # Extract the number and print it
                if number > latest_id:
                    max_file = file if not max_file or number > max_number else max_file
                    max_number = number if not max_file or number > max_number else max_number

            else:
                logging.info("No number found after '-' in the filename.")
        except ValueError:
            continue

    #  Read and print the newest data
    if max_file:
        excel_path = os.path.join(folder_path, max_file)
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        # from the latest ID + 2
        start_row = latest_id + 2
        if start_row <= sheet.max_row:
            for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, values_only=True):
                # Get the header row
                sheet_header = [cell.value for cell in sheet[1]]
                # Format the row data
                formatted_row = []
                for cell in row:
                    if isinstance(cell, datetime.datetime):
                        formatted_row.append(cell.strftime('%Y-%m-%d %H:%M:%S'))  # 格式化日期和时间
                    else:
                        formatted_row.append(cell)

                row = dict(zip(sheet_header, formatted_row))
                logging.info(row)
                rows.append(row)
        else:
            logging.info(f"No new data found starting from row {start_row}.")

        workbook.close()
    else:
        logging.info("No newer Excel file found with an ID greater than the one in latestID.txt.")

    # Write the latest ID to the latestID.txt file
    latest_id_file = os.path.join(folder_path, 'latestID.txt')
    with open(latest_id_file, 'w') as file:
        file.write(str(max_number))

    # Return the rows
    return rows

def newPostRequest(row, service_id):
    headers = {
        'Content-Type': 'application/json',
        'Authorization': Auth_token_input
    }

    body = {
        "change": {
            "title": row["Please describe the Incident/Request in as much detail as possible"],
            "priority_level": "5",
            "type": row["What is the Business Impact of the Incident?"],
            "service_id": service_id,
            "watches": [
                {
                    "watch": {
                        "watched_by": {
                            "id": get_id(row["Email"]),
                            "first_name": row["Name"].split()[0],
                            "last_name": row["Name"].split()[-1],
                            "customer_link": f"/v1/customers/{get_id(row['Email'])}"
                        }
                    }
                }
            ],
            "additional_values": {
                "implementation_start_date_time": row["Start time"],
                "department": row["Which Department are you in?"],
                "system": row["Which System is effected?"],
                "benefit": row["What Benefit Category does the Enhancement or New Capability best align to?"],
                "time_to_benefit": row[
                    "When will Smartshares start receiving the Benefit of the Enhancement or New Capability?"],
                "impact": row["What is the Business Impact of the Enhancement or New Capability?"],
                "strategic_pillar": row["What is the Strategic Alignment of the Enhancement or New Capability?"],
                "reason_for_change": row[
                    "Please describe the Enhancement or New Capability in as much detail as possible."]
            }
        }
    }

    # send POST request with timeout
    response = requests.post('https://deskapi.gotoassist.com/v1/changes', headers=headers, json=body, timeout=10)

    # log the response status code
    if response.status_code == 200:
        logging.info('POST request was successful.')
    else:
        logging.info(f'POST request failed with status code {response.status_code}.')


# read id_email_map.json
def get_id(email):
    return id_email_map.get(email)


# When the value of "Are you reporting an Incident or making a Request?" starts with "Incident", return "IT Support", then get the id from services.json,
# When the value of "Are you reporting an Incident or making a Request?" starts with "Request", return "Smartshares", then get the id from services.json.
def getServicesID(reportingValue):
    if reportingValue.startswith('Incident'):
        return servicesName_id_map.get('IT Support')
    elif reportingValue.startswith('Request'):
        return servicesName_id_map.get('Smartshares')
    else:
        return None


# main
if __name__ == '__main__':
    folder_path = '/Users/tanjunjie/Documents/Forms 365'
    rows = getNewRows(folder_path)
    for row in rows:
        logging.info('Row ID: %s', row["ID"])
        service_id = getServicesID(row["Are you reporting an Incident or making a Request?"])
        newPostRequest(row, service_id)
